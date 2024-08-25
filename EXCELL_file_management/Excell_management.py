from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def add_rows_to_excell_file_openpyxl_2(Filename: str, rows, columns):

    if rows:
        # Optionally, define the column headers if they are not included

        # Convert the list of tuples to a pandas DataFrame
        df = pd.DataFrame(rows, columns=columns)

        # Write the DataFrame to an Excel file
        if not df.empty:
            #file_name = Filename
            today_date = datetime.now().strftime("%Y-%m-%d")
            start_row = int()

            try:
                # Load the existing workbook
                workbook = load_workbook(Filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                with pd.ExcelWriter(Filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=today_date, index=False)
                print(f"File created successfully: {Filename}")


            if today_date in workbook.sheetnames:
                sheet = workbook[today_date]
                # Find the next available row and add spacing
                start_row = sheet.max_row + 2

                # Convert DataFrame to rows
                rows = dataframe_to_rows(df, index=False, header=True)

                # Append the DataFrame to the sheet
                for r_idx, row in enumerate(rows, start_row):
                    for c_idx, value in enumerate(row, 1):
                        sheet.cell(row=r_idx + 1, column=c_idx, value=value)
            else:
                # If the sheet does not exist, create it
                with pd.ExcelWriter(Filename, engine='openpyxl', mode='a') as writer:
                    df.to_excel(writer, sheet_name=today_date, index=False)

                # Save the workbook
            workbook.save(Filename)
            print(f"Data appended successfully to {Filename}")

